import os
import glob
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openai import OpenAI

def read_system_prompt(file_path):
    """从txt文件中读取系统prompt"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except FileNotFoundError:
        print(f"错误：找不到prompt文件 {file_path}")
        return ""
    except Exception as e:
        print(f"读取prompt文件时出错: {e}")
        return ""

def infer(system_prompt, user_query, token=None, model="deepseek-chat"):
    """调用大模型API进行推理"""
    if not token:
        print("错误：缺少API token")
        return ""
    
    client = OpenAI(api_key=token, base_url="https://api.deepseek.com/v1")
    
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_query},
            ],
            stream=False
        )
        
        result = response.choices[0].message.content
        return result.strip()
    
    except Exception as e:
        print(f"调用API出错: {e}")
        return ""

def save_excel(save_path, title_list, content_list):
    """保存结果到Excel文件"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(title_list)
        
        for content in content_list:
            ws.append([content])
        
        # 确保目录存在
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        
        # 如果文件已存在则覆盖
        if os.path.exists(save_path):
            os.remove(save_path)
            
        wb.save(save_path)
        wb.close()
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")

def get_training_examples(annotation_file):
    """从标注文件中读取训练示例"""
    examples = []
    
    try:
        wb = openpyxl.load_workbook(annotation_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):  # 跳过标题行
            if len(row) >= 4:  # 确保有至少4列
                query = str(row[2].value) if row[2].value else ""
                label = str(row[3].value) if row[3].value else ""
                
                if query and label:
                    examples.append((query, label))
        
        wb.close()
    except Exception as e:
        print(f"读取标注文件时出错: {e}")
    
    return examples

def create_enhanced_prompt(base_prompt, examples):
    """创建增强后的prompt"""
    if not examples:
        return base_prompt
    
    example_section = "\n\n已标注示例（格式：'query' -> label）：\n"
    for query, label in examples:
        example_section += f'"{query}" -> {label}\n'
    
    return base_prompt + example_section

def process_single_file(input_file, enhanced_prompt, token, model):
    """处理单个Excel文件"""
    results = []
    
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):  # 从第二行开始
            query = ""
            
            for cell in row:
                col_letter = get_column_letter(cell.column)
                if col_letter == 'D':  # 假设query在第4列(D列)
                    query = str(cell.value) if cell.value else ""
                    break
            
            if query:
                response = infer(enhanced_prompt, query, token, model)
                results.append(response)
                print(f"处理: {query[:50]}... -> {response}")
        
        wb.close()
    except Exception as e:
        print(f"处理文件 {input_file} 时出错: {e}")
    
    return results

def main(source_files, annotation_file, prompt_file, token, model, output_dir):
    """主处理函数"""
    # 验证输入
    if not source_files:
        print("错误：没有找到输入文件")
        return
    
    # 读取系统prompt
    system_prompt = read_system_prompt(prompt_file)
    if not system_prompt:
        return
    
    # 读取训练示例
    training_examples = get_training_examples(annotation_file)
    
    # 创建增强prompt
    enhanced_prompt = create_enhanced_prompt(system_prompt, training_examples)
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    # 处理每个文件
    total_files = len(source_files)
    print(f"开始处理 {total_files} 个文件...")
    
    for i, input_file in enumerate(source_files, 1):
        print(f"\n[{i}/{total_files}] 正在处理: {input_file}")
        
        # 处理当前文件
        results = process_single_file(input_file, enhanced_prompt, token, model)
        
        # 生成输出路径
        base_name = os.path.basename(input_file)
        output_path = os.path.join(output_dir, f"classified_{base_name}")
        
        # 保存结果
        save_excel(output_path, ["模型判断结果"], results)
        print(f"结果已保存到: {output_path}")
    
    print("\n所有文件处理完成！")

if __name__ == "__main__":
    # 配置参数
    INPUT_DIR = "input_files"  # 输入文件目录
    ANNOTATION_FILE = "labeled_data.xlsx"  # 标注文件路径
    PROMPT_FILE = "system_prompt.txt"  # prompt文件路径
    API_TOKEN = "sk-你的API密钥"  # 替换为你的实际API密钥
    MODEL_NAME = "deepseek-chat"  # 模型名称
    OUTPUT_DIR = "classification_results"  # 输出目录
    
    # 获取所有输入文件
    source_files = glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    
    # 运行主程序
    main(
        source_files=source_files,
        annotation_file=ANNOTATION_FILE,
        prompt_file=PROMPT_FILE,
        token=API_TOKEN,
        model=MODEL_NAME,
        output_dir=OUTPUT_DIR
    )