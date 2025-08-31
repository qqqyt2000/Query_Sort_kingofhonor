import requests
import json
import os
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter

def infer(system_prompt, user_query, token=None, model=None):
    # 定义接口域名
    base_url = "http://hunyuanapi.woa.com"  # 替换为实际域名

    # 示例2：POST 请求（提交JSON数据）
    post_url = base_url + "/openapi/v1/chat/completions"
    headers = {
        "Content-Type": "json",
        "Authorization": "Bearer " + token if token else "",
        }
    data = {"model": model, "messages": [{"role": "system", "content": system_prompt},
                                           {"role": "user", "content": user_query}]}

    response = requests.post(url=post_url, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        result = response.json()
        result = result["choices"][0]["message"]["content"]
        if result[-1] == '\n':
            result = result[:-1]
        return result
    return ""


def save_excel(save_path, title_list, content_list):
    wb_trg = Workbook()
    ws_trg = wb_trg.active
    ws_trg.append(title_list)
    for content in content_list:
        ws_trg.append([content])
    if os.path.exists(save_path):
        os.remove(save_path)
    wb_trg.save(save_path)
    wb_trg.close()


def main(source_file, system_prompt, token=None, model=None, save_path=None):
    wb_source = openpyxl.load_workbook(source_file)
    ws_source = wb_source.active  # 获取活动工作表

    # add data
    data = []
    for row in ws_source.iter_rows():  # 遍历每一行
        query = ""
        response = ""

        for cell in row:  # 遍历行中的每个单元格
            # 获取行列序号（注意openpyxl行列索引从1开始）
            row_index = cell.row
            col_index = cell.column
            col_letter = get_column_letter(col_index)  # 列字母编号（如'D'）
            
            if col_letter == 'D':
                query = cell.value if isinstance(cell.value, str) else ""
            
        response = infer(system_prompt, query, token, model) if query else ""
        if row_index > 1:
            print(f"{query} -> {response}")
            data.append(response)
    title_list = ["回复"]
    save_excel(save_path, title_list, data)
    return

if __name__ == "__main__":
    # prompt
    system_prompt = """
    你是一个判断moba游戏玩家语音情绪意图语义的AI专家，请判断以下句子的语义是：1、消极且明确表达AI不听指挥；2、其他消极言论；3、中性；4、积极，直接输出数字，不要输出其他内容。

其中，将以下几类内容定义为“消极且明确表达AI不听指挥”：
1. 直接指责型
句式特点：明确点出对方未执行指令或行为错误。
• “你为什么不听我的？”
• “我说了XXX，你没听见吗？”
• 例：“我说了辅助跟我，你没听见吗？”
• “你非要XXX是吧？”（强调故意违抗）
• 例：“你非要抢我蓝是吧？”

2. 抱怨对比型
句式特点：通过对比“预期”和“现实”表达不满。
• “让你XXX（你偏要XXX）”
• 例：“让你推塔，你偏要打野。”
• “我说了XXX（结果你XXX）”
• 例：“我说了集合，结果你一个人去送。”

3. 反问质问型
句式特点：用反问句强化指责，隐含“你应该听我的”。
• “你是听不懂人话吗？”
• “我说话是放屁吗？”（粗俗但常见）
• “你眼睛长哪儿了？”（指责未注意指令）

4. 消极威胁型
句式特点：用消极行为或威胁迫使对方服从。
• “再不XXX我就XXX。”
• 例：“再不跟我，我就挂机。”
• “随便吧，你们爱怎么玩怎么玩。”（放弃沟通）

5. 无奈总结型
句式特点：表达对队友“不听话”的无力感。
• “带不动，一群菜鸟。”
• “这队友没救了。”
• “你们开心就好。”（反讽）

6. 重复强调型
句式特点：因未被响应而重复指令，隐含不满。可以通过
• “给我蓝！给我蓝！给我蓝！”
“撤退！撤退！听不懂吗？

    """
    # 输入excel路径
    source_file = "output.xlsx"
    # hunyuan token
    token = "5GnAlrgdCrQtZlcJAergnxOcDUsatAKD"
    # 选择的模型
    model = "hunyuan-t1-latest"
    # 输出excel路径
    save_path = "output_LLM_review.xlsx"
    
    main(source_file, system_prompt, token, model, save_path)