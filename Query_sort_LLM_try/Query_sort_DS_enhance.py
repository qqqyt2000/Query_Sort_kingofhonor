import requests
from openai import OpenAI
import json
import os
from openpyxl import Workbook
import openpyxl
from openpyxl.utils import get_column_letter
import logging
from typing import List, Tuple, Optional
import time

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class QueryClassifier:
    """查询分类器类"""
    
    def __init__(self, api_key: str, base_url: str = "https://api.deepseek.com/v1", model: str = "deepseek-chat"):
        self.client = OpenAI(api_key=api_key, base_url=base_url)
        self.model = model
        
    def read_system_prompt(self, file_path: str) -> str:
        """从txt文件中读取系统prompt"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
        except FileNotFoundError:
            logger.error(f"系统prompt文件不存在: {file_path}")
            raise
        except Exception as e:
            logger.error(f"读取系统prompt文件失败: {e}")
            raise
    
    def infer(self, system_prompt: str, user_query: str, max_retries: int = 3) -> str:
        """调用API进行推理，支持重试机制"""
        for attempt in range(max_retries):
            try:
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_query},
                    ],
                    stream=False,
                    timeout=30
                )
                
                result = response.choices[0].message.content
                return result.strip() if result else ""
                
            except Exception as e:
                logger.warning(f"API调用失败 (尝试 {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # 指数退避
                else:
                    logger.error(f"API调用最终失败: {e}")
                    return ""
    
    def get_training_examples(self, annotation_file: str, query_col: str = 'C', label_col: str = 'D') -> List[Tuple[str, str]]:
        """从标注文件中读取训练示例"""
        try:
            wb = openpyxl.load_workbook(annotation_file)
            ws = wb.active
            examples = []
            
            query_col_idx = openpyxl.utils.column_index_from_string(query_col)
            label_col_idx = openpyxl.utils.column_index_from_string(label_col)
            
            for row in ws.iter_rows(min_row=2):  # 跳过标题行
                query = row[query_col_idx - 1].value if row[query_col_idx - 1].value else ""
                label = row[label_col_idx - 1].value if row[label_col_idx - 1].value else ""
                if query and label:
                    examples.append((str(query).strip(), str(label).strip()))
            
            logger.info(f"成功读取 {len(examples)} 个训练示例")
            return examples
            
        except Exception as e:
            logger.error(f"读取训练示例失败: {e}")
            return []
    
    def create_enhanced_prompt(self, base_prompt: str, examples: List[Tuple[str, str]]) -> str:
        """增强prompt，添加训练示例"""
        if not examples:
            logger.warning("没有训练示例，使用原始prompt")
            return base_prompt
            
        example_section = "\n\n下面是一些标注好的示例（格式为'query -> label'），请参考这些示例进行判断：\n"
        for query, label in examples:
            example_section += f'"{query}" -> {label}\n'
        
        return base_prompt + example_section
    
    def process_queries(self, source_file: str, query_col: str = 'F', enhanced_prompt: str = "") -> List[str]:
        """处理查询文件"""
        try:
            wb_source = openpyxl.load_workbook(source_file)
            ws_source = wb_source.active
            results = []
            
            query_col_idx = openpyxl.utils.column_index_from_string(query_col)
            
            for row_idx, row in enumerate(ws_source.iter_rows(), 1):
                if row_idx == 1:  # 跳过标题行
                    continue
                    
                query = row[query_col_idx - 1].value if row[query_col_idx - 1].value else ""
                if not query:
                    results.append("")
                    continue
                
                query = str(query).strip()
                logger.info(f"处理查询 {row_idx}: {query[:50]}...")
                
                response = self.infer(enhanced_prompt, query)
                results.append(response)
                
                # 添加延迟避免API限制
                time.sleep(0.5)
            
            return results
            
        except Exception as e:
            logger.error(f"处理查询文件失败: {e}")
            return []
    
    def save_results(self, save_path: str, results: List[str], title: str = "回复"):
        """保存结果到Excel文件"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append([title])
            
            for result in results:
                ws.append([result])
            
            # 确保目录存在
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            wb.save(save_path)
            wb.close()
            logger.info(f"结果已保存到: {save_path}")
            
        except Exception as e:
            logger.error(f"保存结果失败: {e}")
            raise

def main():
    """主函数"""
    # 配置参数
    config = {
        'source_file': "select_0729.xlsx",
        'annotation_file': "label.xlsx", 
        'prompt_file': "system_prompt.txt",
        'token': "sk-9c97270e0e9e4d0cb0c296417e5b4232",
        'model': "deepseek-chat",
        'save_path': "select_0729_classification_results.xlsx",
        'query_col': 'F',  # 查询所在列
        'query_col_annotation': 'C',  # 标注文件中查询列
        'label_col_annotation': 'D'   # 标注文件中标签列
    }
    
    try:
        # 初始化分类器
        classifier = QueryClassifier(config['token'], model=config['model'])
        
        # 读取系统prompt
        system_prompt = classifier.read_system_prompt(config['prompt_file'])
        
        # 读取训练示例
        training_examples = classifier.get_training_examples(
            config['annotation_file'], 
            config['query_col_annotation'], 
            config['label_col_annotation']
        )
        
        # 增强系统prompt
        enhanced_prompt = classifier.create_enhanced_prompt(system_prompt, training_examples)
        
        # 处理查询
        results = classifier.process_queries(
            config['source_file'], 
            config['query_col'], 
            enhanced_prompt
        )
        
        # 保存结果
        classifier.save_results(config['save_path'], results)
        
        logger.info("处理完成！")
        
    except Exception as e:
        logger.error(f"程序执行失败: {e}")
        raise

if __name__ == "__main__":
    main()